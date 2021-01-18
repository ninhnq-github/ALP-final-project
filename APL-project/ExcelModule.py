import pandas as pd
from openpyxl import Workbook
from openpyxl import load_workbook
from dbConnect import *

def ExcelFile(path, subject='None'): #Nhập dữ liệu file excel ở địa chỉ path
    xl = pd.ExcelFile(path)
    df = pd.read_excel(xl, 0, header=None)
    table = []
    for i in range(0,len(df.iloc[:, 0])-6):
        table.append([])
        table[i].append(df.at[i+6, 1])
        table[i].append(str(df.at[i+6, 2]) + str(df.at[i+6, 3]))
        table[i].append(subject)
        table[i].append('0')
    return table

#ExcelFile('D:\CODE\PYTHON\ALP-final-project\APL-project\\201LLCT120205_05.xls')

def Export(path): #xuất dữ liệu sang file excel ở địa chỉ path
    book = Workbook()
    sheet = book.active    
    sheet['A1'] = 'MSSV'
    sheet['B1'] = 'Họ tên'
    sheet['C1'] = 'Môn thi'
    sheet['D1'] = 'Điểm'
    table = dbConnect.getStudent()
    for i in range(len(table)):
        sheet['A' + str(i + 2)] = table[i][0]
        sheet['B' + str(i + 2)] = table[i][1]
        sheet['C' + str(i + 2)] = table[i][2]
        sheet['D' + str(i + 2)] = table[i][3]
        book.save(path + "/BangDiem.xlsx")

def ExportToExcel(input_path,export_path):
    workbook = load_workbook(filename=input_path)
    sheet = workbook.active
    IDlist = []
    table = getStudent()
    for i in range(len(table)):
        IDlist.append(table[i][0])
    for i in range(6,len(sheet['A'])):
        try:
            #print(sheet['B'+str(i+1)].value)
            index = IDlist.index(sheet['B'+str(i+1)].value)
            sheet['G'+str(i+1)].value = table[index][3]
        except:
            print('lost at :',sheet['B'+str(i+1)].value)
            return False
    workbook.save(filename=export_path)
    return True

#ExcelFile('D:\CODE\PYTHON\ALP-final-project\APL-project\STDList.xlsx','Math')
#ExportToExcel('D:\CODE\PYTHON\ALP-final-project\APL-project\\201LLCT120205_05.xlsx','D:\CODE\PYTHON\ALP-final-project\APL-project\\DIEM201LLCT120205_05.xlsx')

def importExcelData(path, subject):
    table = ExcelFile(path, subject)
    for row in table:
        if (str(row[0])!='nan'):
            print(insertStudent(row[0],row[1],row[2],row[3]))

def fromExcelToCsv(path,csvname):
    file = open(csvname,"w", encoding="utf-8")
    table = ExcelFile(path, 'Math')
    for row in table:
        if (str(row[0])!='nan'):
            #print('%s,%s,%s,%s'%(row[0],row[1],row[2],row[3]))
            file.write('%s,%s,%s,%s\n'%(row[0],row[1],row[2],row[3]))
    file.close()

#importExcelData('D:\CODE\PYTHON\ALP-final-project\APL-project\STDList.xlsx','Math')
#fromExcelToCsv('D:\CODE\PYTHON\ALP-final-project\APL-project\STDList.xlsx','D:\CODE\PYTHON\ALP-final-project\APL-project\StudentList.csv')

#import_csv_to_mysql(filename)